"""Utilidades compartidas para importación/exportación batch."""
import csv
import json
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import send_file


# ── Parsers ──────────────────────────────────────────────────────────────────

def parse_file(file):
    """Parsea un archivo CSV, JSON o XLSX y retorna lista de dicts con claves en minúscula."""
    name = file.filename.lower()
    if name.endswith('.csv'):
        return _parse_csv(file)
    elif name.endswith('.json'):
        return _parse_json(file)
    elif name.endswith('.xlsx'):
        return _parse_xlsx(file)
    raise ValueError('Formato no soportado. Use CSV, JSON o XLSX')


def _parse_csv(file):
    content = file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(content))
    return [{k.strip().lower(): v for k, v in row.items()} for row in reader]


def _parse_json(file):
    data = json.load(file)
    if not isinstance(data, list):
        raise ValueError('El JSON debe ser un array de objetos')
    return [{k.lower(): v for k, v in row.items()} for row in data]


def _parse_xlsx(file):
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        obj = {headers[i]: (str(v).strip() if v is not None else '')
               for i, v in enumerate(row) if i < len(headers) and headers[i]}
        if any(obj.values()):
            result.append(obj)
    return result


# ── Helpers de valor ─────────────────────────────────────────────────────────

def int_or_none(val):
    v = str(val).strip() if val else ''
    return int(float(v)) if v else None


def float_or_zero(val):
    v = str(val).strip() if val else ''
    return float(v) if v else 0.0


def bool_col(val):
    """'1', 'true', 'si', 'yes' → 1; resto → 0."""
    return 1 if str(val).strip().lower() in ('1', 'true', 'si', 'sí', 'yes') else 0


# ── Exportadores ─────────────────────────────────────────────────────────────

def export_csv(rows, campos, filename):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=campos, extrasaction='ignore')
    writer.writeheader()
    for r in rows:
        writer.writerow({k: (r.get(k, '') if r.get(k) is not None else '') for k in campos})
    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return send_file(out, mimetype='text/csv', as_attachment=True, download_name=filename)


def export_json(rows, campos, filename):
    data = [{k: (r.get(k, '') if r.get(k) is not None else '') for k in campos} for r in rows]
    out = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2, default=str).encode('utf-8'))
    return send_file(out, mimetype='application/json', as_attachment=True, download_name=filename)


def export_xlsx(rows, campos, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename.replace('.xlsx', '')[:31]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2980B9')

    ws.append(campos)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([(r.get(k, '') if r.get(k) is not None else '') for k in campos])

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def plantilla_csv(campos, ejemplo, filename):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(campos)
    writer.writerow(ejemplo)
    out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
    return send_file(out, mimetype='text/csv', as_attachment=True, download_name=filename)


def plantilla_json(campos, ejemplo, filename):
    data = [dict(zip(campos, ejemplo))]
    out = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
    return send_file(out, mimetype='application/json', as_attachment=True, download_name=filename)


def plantilla_xlsx(campos, ejemplo, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Plantilla'
    ws.append(campos)
    ws.append(ejemplo)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
