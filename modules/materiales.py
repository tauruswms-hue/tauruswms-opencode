from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, session
import os
import csv
import json
import io
from dotenv import load_dotenv
import openpyxl
from modules.db_config import get_db_connection, _get_admin_connection
from modules.sql_dialect import cast_as_char, execute_insert

load_dotenv()
materiales_bp = Blueprint('materiales', __name__)

PICKING_METODOS_LABELS = {
    'fifo': 'FIFO (Primero en entrar, primero en salir)',
    'lifo': 'LIFO (Último en entrar, primero en salir)',
    'fefo': 'FEFO (Próximo a vencer)',
    'libre': 'Picking Libre',
}


def _get_picking_metodos(tenant_id):
    """Devuelve la lista de métodos de picking habilitados para el tenant."""
    if not tenant_id:
        return list(PICKING_METODOS_LABELS.keys())
    conn = _get_admin_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT metodosdepicking FROM tenants WHERE id = %s", (tenant_id,))
            row = cursor.fetchone()
    finally:
        conn.close()
    if not row or not row.get('metodosdepicking'):
        return list(PICKING_METODOS_LABELS.keys())
    try:
        metodos = json.loads(row['metodosdepicking'])
    except Exception:
        metodos = row['metodosdepicking']
    if isinstance(metodos, str):
        metodos = [metodos]
    metodos = [m for m in metodos if m in PICKING_METODOS_LABELS]
    return metodos or list(PICKING_METODOS_LABELS.keys())


def _get_picking_metodo_default(tenant_id):
    """Devuelve el método de picking por defecto configurado para el tenant."""
    if not tenant_id:
        return 'libre'
    conn = _get_admin_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT metodo_picking_default FROM tenants WHERE id = %s", (tenant_id,))
            row = cursor.fetchone()
    finally:
        conn.close()
    default = (row.get('metodo_picking_default') or 'libre') if row else 'libre'
    return default if default in PICKING_METODOS_LABELS else 'libre'


def _metodo_picking_valido(metodo, default='libre'):
    if metodo in PICKING_METODOS_LABELS:
        return metodo
    return default if default in PICKING_METODOS_LABELS else 'libre'


def validar_ean(barcode):
    if not barcode or not barcode.strip():
        return True, None
    barcode = barcode.strip()
    if not barcode.isdigit() or len(barcode) not in (8, 13):
        return False, 'El código de barras debe tener 8 (EAN-8) o 13 (EAN-13) dígitos numéricos.'
    es_ean8 = len(barcode) == 8
    digitos = [int(d) for d in barcode]
    suma = 0
    for i in range(len(digitos) - 1):
        peso = 3 if (i % 2 == 0 and es_ean8) or (i % 2 != 0 and not es_ean8) else 1
        suma += digitos[i] * peso
    check_calculado = (10 - (suma % 10)) % 10
    if check_calculado != digitos[-1]:
        return False, 'El código de barras tiene un dígito verificador inválido.'
    return True, None


def validar_gtin14(barcode):
    if not barcode or not barcode.strip():
        return True, None
    barcode = barcode.strip()
    if not barcode.isdigit() or len(barcode) != 14:
        return False, 'El GTIN-14 debe tener exactamente 14 dígitos numéricos.'
    digitos = [int(d) for d in barcode]
    suma = 0
    for i in range(13):
        peso = 3 if i % 2 == 0 else 1
        suma += digitos[i] * peso
    check_calculado = (10 - (suma % 10)) % 10
    if check_calculado != digitos[13]:
        return False, 'El GTIN-14 tiene un dígito verificador inválido.'
    return True, None


def get_tenant_filter():
    return session.get('tenant_id')


@materiales_bp.route('/materiales')
def listar():
    tenant_id = get_tenant_filter()
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            sql_mat = """
                SELECT m.*, c.nombre as categoria_nombre,
                       p.razonsocial as proveedor_habitual,
                       u.nombre as unidad_medida_nombre, u.simbolo as unidad_medida_simbolo
                FROM materiales m
                LEFT JOIN categorias c ON m.categoria_id = c.id_categoria AND (%s IS NULL OR c.tenant_id = %s)
                LEFT JOIN material_proveedor mp ON mp.id_material = m.id AND mp.es_habitual = 1 AND (%s IS NULL OR mp.tenant_id = %s)
                LEFT JOIN proveedores p ON p.id = mp.id_proveedor AND (%s IS NULL OR p.tenant_id = %s)
                LEFT JOIN unidades_medida u ON m.unidad_medida_id = u.id_unidad AND (%s IS NULL OR u.tenant_id = %s)
                WHERE (%s IS NULL OR m.tenant_id = %s)
                ORDER BY m.id DESC
            """
            cursor.execute(sql_mat, (tenant_id, tenant_id, tenant_id, tenant_id, tenant_id, tenant_id, tenant_id, tenant_id, tenant_id, tenant_id))
            materiales = cursor.fetchall()

            cursor.execute("SELECT * FROM categorias WHERE activo = 1 AND (%s IS NULL OR tenant_id = %s)", (tenant_id, tenant_id))
            categorias = cursor.fetchall()

            cursor.execute("SELECT id, razonsocial FROM proveedores WHERE activo = 1 AND (%s IS NULL OR tenant_id = %s)", (tenant_id, tenant_id))
            proveedores = cursor.fetchall()

            cursor.execute("SELECT id_unidad, codigo, nombre, simbolo FROM unidades_medida WHERE activo = 1 AND (%s IS NULL OR tenant_id = %s) ORDER BY nombre", (tenant_id, tenant_id))
            unidades = cursor.fetchall()

            cursor.execute("SELECT mp.* FROM material_proveedor mp WHERE %s IS NULL OR mp.tenant_id = %s", (tenant_id, tenant_id))
            relaciones = cursor.fetchall()

            cursor.execute(f"""
                SELECT id, id_material, nombre, codigo_barras,
                       {cast_as_char('cantidad_unidades')} AS cantidad_unidades,
                       peso_bruto, peso_neto
                FROM material_presentaciones
                WHERE activo = 1 AND (%s IS NULL OR tenant_id = %s)
                ORDER BY id_material, cantidad_unidades
            """, (tenant_id, tenant_id))
            presentaciones = cursor.fetchall()

        return render_template('materiales.html',
                               materiales=materiales,
                               categorias=categorias,
                               proveedores=proveedores,
                               relaciones=relaciones,
                               presentaciones=presentaciones,
                               unidades=unidades,
                               picking_metodos=[{'value': m, 'label': PICKING_METODOS_LABELS[m]} for m in _get_picking_metodos(tenant_id)],
                               picking_metodo_default=_get_picking_metodo_default(tenant_id),
                               picking_labels=PICKING_METODOS_LABELS)
    finally:
        conn.close()


@materiales_bp.route('/materiales/guardar', methods=['POST'])
def guardar():
    d = request.form
    m_id = d.get('id')

    prov_ids = request.form.getlist('prov_ids[]')
    prov_codigos = request.form.getlist('prov_codigos[]')
    prov_habitual = d.get('prov_habitual', None)

    pres_nombres = request.form.getlist('pres_nombres[]')
    pres_barcodes = request.form.getlist('pres_barcodes[]')
    pres_cantidades = request.form.getlist('pres_cantidades[]')
    pres_pesos_brutos = request.form.getlist('pres_pesos_brutos[]')
    pres_pesos_netos = request.form.getlist('pres_pesos_netos[]')

    import json
    
    barcode = d.get('codigo_barras', '')
    valido, error_msg = validar_ean(barcode)
    if not valido:
        flash(error_msg or 'Código de barras inválido', 'danger')
        return redirect(url_for('materiales.listar'))

    for i, gtin in enumerate(pres_barcodes):
        if gtin and gtin.strip():
            valido_gtin, error_gtin = validar_gtin14(gtin.strip())
            if not valido_gtin:
                nombre_p = pres_nombres[i].strip() if i < len(pres_nombres) and pres_nombres[i] else f'Presentación {i+1}'
                flash(f'Presentación "{nombre_p}": {error_gtin}', 'danger')
                return redirect(url_for('materiales.listar'))

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            tenant_id = get_tenant_filter()
            trazabilidad = d.get('trazabilidad', 'ninguna')
            stock_min = float(d.get('stock_minimo') or 0)
            stock_max = float(d.get('stock_maximo') or 0)
            peso_bruto = float(d.get('peso_bruto') or 0) or None
            peso_neto = float(d.get('peso_neto') or 0) or None
            categoria_id = int(d.get('categoria_id') or 0) or None
            unidad_id = int(d.get('unidad_medida_id') or 0) or None
            metodo_picking = _metodo_picking_valido((d.get('metodo_picking') or '').strip().lower(),
                                                    _get_picking_metodo_default(tenant_id))

            if m_id:
                cursor.execute("""
                    UPDATE materiales SET
                        codigo = %s, nombre = %s, descripcion = %s, codigo_barras = %s,
                        categoria_id = %s, stock_minimo = %s, stock_maximo = %s,
                        unidad_medida_id = %s, trazabilidad = %s, metodo_picking = %s,
                        peso_bruto = %s, peso_neto = %s
                    WHERE id = %s AND (%s IS NULL OR tenant_id = %s)
                """, (d.get('codigo'), d.get('nombre'), d.get('descripcion') or '', 
                      barcode or None, categoria_id, stock_min, stock_max,
                      unidad_id, trazabilidad, metodo_picking, peso_bruto, peso_neto,
                      m_id, tenant_id, tenant_id))
                current_id = int(m_id)
            else:
                current_id = execute_insert(cursor, """
                    INSERT INTO materiales (codigo, nombre, descripcion, codigo_barras, categoria_id,
                        stock_minimo, stock_maximo, unidad_medida_id, trazabilidad, metodo_picking,
                        peso_bruto, peso_neto, tenant_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (d.get('codigo'), d.get('nombre'), d.get('descripcion') or '', 
                      barcode or None, categoria_id, stock_min, stock_max,
                      unidad_id, trazabilidad, metodo_picking, peso_bruto, peso_neto, tenant_id))

            cursor.execute("DELETE FROM material_proveedor WHERE id_material = %s AND (%s IS NULL OR tenant_id = %s)", (current_id, tenant_id, tenant_id))
            for i, prov_id in enumerate(prov_ids):
                if prov_id:
                    codigo_prov = prov_codigos[i] if i < len(prov_codigos) else ''
                    es_habitual = 1 if prov_habitual is not None and int(prov_habitual) == i else 0
                    cursor.execute("""
                        INSERT INTO material_proveedor (id_material, id_proveedor, codigo_referencia_prov, es_habitual, tenant_id)
                        VALUES (%s, %s, %s, %s, %s)
                    """, (current_id, prov_id, codigo_prov, es_habitual, tenant_id))

            cursor.execute("DELETE FROM material_presentaciones WHERE id_material = %s AND (%s IS NULL OR tenant_id = %s)", (current_id, tenant_id, tenant_id))
            for i, nombre in enumerate(pres_nombres):
                if nombre and nombre.strip():
                    barcode_pres = pres_barcodes[i].strip() if i < len(pres_barcodes) and pres_barcodes[i] else ''
                    cantidad = float(pres_cantidades[i]) if i < len(pres_cantidades) and pres_cantidades[i] else 1.0
                    pres_pb = float(pres_pesos_brutos[i]) if i < len(pres_pesos_brutos) and pres_pesos_brutos[i] else None
                    pres_pn = float(pres_pesos_netos[i]) if i < len(pres_pesos_netos) and pres_pesos_netos[i] else None
                    cursor.execute("""
                        INSERT INTO material_presentaciones (id_material, nombre, codigo_barras, cantidad_unidades, peso_bruto, peso_neto, tenant_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (current_id, nombre.strip(), barcode_pres or None, cantidad, pres_pb, pres_pn, tenant_id))

            conn.commit()
            flash("Material guardado correctamente", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error: {str(e)}", "danger")
    finally:
        conn.close()
    return redirect(url_for('materiales.listar'))


def _parse_csv(file):
    content = file.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(content))
    return [{k.strip().lower(): v for k, v in row.items()} for row in reader]


def _parse_json(file):
    data = json.load(file)
    if not isinstance(data, list):
        raise ValueError('El JSON debe ser un array de objetos')
    return [{k.lower(): v for k, v in row.items()} for row in data]


def _parse_xlsx(file, hoja=None):
    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    if hoja:
        if hoja not in wb.sheetnames:
            raise ValueError(
                f'La hoja "{hoja}" no existe en el archivo. '
                f'Hojas disponibles: {", ".join(wb.sheetnames)}'
            )
        ws = wb[hoja]
    else:
        ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    result = []
    for row in rows[1:]:
        obj = {headers[i]: (str(v).strip() if v is not None else '') for i, v in enumerate(row) if i < len(headers) and headers[i]}
        if any(obj.values()):
            result.append(obj)
    return result


@materiales_bp.route('/materiales/importar', methods=['POST'])
def importar():
    file = request.files.get('archivo')
    if not file or file.filename == '':
        return jsonify({'error': 'No se proporcionó archivo'}), 400

    tenant_id = session.get('tenant_id')
    if not tenant_id:
        return jsonify({'error': 'No se encontró el tenant del usuario'}), 400

    filename = file.filename.lower()
    try:
        if filename.endswith('.csv'):
            rows = _parse_csv(file)
        elif filename.endswith('.json'):
            rows = _parse_json(file)
        elif filename.endswith('.xlsx'):
            rows = _parse_xlsx(file, request.form.get('hoja'))
        else:
            return jsonify({'error': 'Formato no soportado. Use CSV, JSON o XLSX'}), 400
    except Exception as e:
        return jsonify({'error': f'Error al leer el archivo: {str(e)}'}), 400

    insertados = 0
    omitidos = []
    errores = []

    metodo_default = _get_picking_metodo_default(tenant_id)

    conn = get_db_connection()
    try:
        for i, row in enumerate(rows, start=1):
            codigo = str(row.get('codigo', '') or '').strip()
            nombre = str(row.get('nombre', '') or '').strip()
            if not codigo or not nombre:
                errores.append({'fila': i, 'codigo': codigo or '(vacío)', 'razon': 'Código y Nombre son obligatorios'})
                continue
            try:
                with conn.cursor() as cursor:
                    cursor.execute("SELECT id FROM materiales WHERE codigo = %s AND tenant_id = %s", (codigo, tenant_id))
                    if cursor.fetchone():
                        omitidos.append(codigo)
                        continue

                    def _int_or_none(val):
                        v = str(val).strip() if val else ''
                        return int(float(v)) if v else None

                    def _float_or_zero(val):
                        v = str(val).strip() if val else ''
                        return float(v) if v else 0.0

                    traz = str(row.get('trazabilidad', '') or '').strip().lower()
                    if traz not in ('lote', 'serie', 'ninguna'):
                        traz = 'ninguna'

                    metodo_picking = _metodo_picking_valido(str(row.get('metodo_picking', '') or '').strip().lower(),
                                                            metodo_default)

                    barcode_import = str(row.get('codigo_barras', '') or '').strip()
                    valido_cb, error_cb = validar_ean(barcode_import)
                    if not valido_cb:
                        errores.append({'fila': i, 'codigo': codigo, 'razon': error_cb})
                        continue

                    if barcode_import:
                        cursor.execute("SELECT id FROM materiales WHERE codigo_barras = %s AND tenant_id = %s", (barcode_import, tenant_id))
                        if cursor.fetchone():
                            errores.append({'fila': i, 'codigo': codigo, 'razon': f'El código de barras {barcode_import} ya está asignado a otro material.'})
                            continue

                    material_id = execute_insert(cursor, """
                        INSERT INTO materiales
                            (codigo, nombre, descripcion, codigo_barras,
                             categoria_id, stock_minimo, stock_maximo,
                             unidad_medida_id, trazabilidad, metodo_picking,
                             peso_bruto, peso_neto,
                             tenant_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        codigo,
                        nombre,
                        str(row.get('descripcion', '') or '').strip() or None,
                        barcode_import or None,
                        _int_or_none(row.get('categoria_id')),
                        _float_or_zero(row.get('stock_minimo')),
                        _float_or_zero(row.get('stock_maximo')),
                        _int_or_none(row.get('unidad_medida_id')),
                        traz,
                        metodo_picking,
                        _float_or_zero(row.get('peso_bruto')) or None,
                        _float_or_zero(row.get('peso_neto')) or None,
                        tenant_id,
                    ))

                    id_prov_hab = _int_or_none(row.get('id_proveedor_habitual'))
                    if id_prov_hab:
                        cursor.execute("SELECT id FROM proveedores WHERE id = %s AND (%s IS NULL OR tenant_id = %s)", (id_prov_hab, tenant_id, tenant_id))
                        if cursor.fetchone():
                            cod_ref_prov = str(row.get('codigo_referencia_prov', '') or '').strip() or None
                            cursor.execute("""
                                INSERT INTO material_proveedor (id_material, id_proveedor, codigo_referencia_prov, es_habitual, tenant_id)
                                VALUES (%s, %s, %s, 1, %s)
                            """, (material_id, id_prov_hab, cod_ref_prov, tenant_id))

                    insertados += 1
            except Exception as e:
                errores.append({'fila': i, 'codigo': codigo, 'razon': str(e)})
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

    return jsonify({'insertados': insertados, 'omitidos': omitidos, 'errores': errores})


@materiales_bp.route('/materiales/exportar/<formato>')
def exportar(formato):
    tenant_id = get_tenant_filter()
    CAMPOS = ['codigo', 'nombre', 'descripcion', 'codigo_barras',
              'categoria_id', 'categoria_nombre', 'stock_minimo', 'stock_maximo',
              'unidad_medida_id', 'unidad_medida_nombre', 'trazabilidad', 'metodo_picking',
              'peso_bruto', 'peso_neto',
              'id_proveedor_habitual', 'proveedor_habitual_nombre', 'codigo_referencia_prov']

    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("""
                SELECT m.codigo, m.nombre, m.descripcion, m.codigo_barras,
                       m.categoria_id, c.nombre AS categoria_nombre,
                       m.stock_minimo, m.stock_maximo,
                       m.unidad_medida_id, u.nombre AS unidad_medida_nombre, m.trazabilidad,
                       m.metodo_picking,
                       m.peso_bruto, m.peso_neto,
                       mp.id_proveedor AS id_proveedor_habitual,
                       p.razonsocial AS proveedor_habitual_nombre,
                       mp.codigo_referencia_prov
                FROM materiales m
                LEFT JOIN categorias c ON m.categoria_id = c.id_categoria
                LEFT JOIN unidades_medida u ON m.unidad_medida_id = u.id_unidad
                LEFT JOIN material_proveedor mp ON mp.id_material = m.id AND mp.es_habitual = 1
                LEFT JOIN proveedores p ON p.id = mp.id_proveedor
                WHERE (%s IS NULL OR m.tenant_id = %s)
                ORDER BY m.codigo
            """, (tenant_id, tenant_id))
            rows = cursor.fetchall()
    finally:
        conn.close()

    if formato == 'csv':
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=CAMPOS, extrasaction='ignore')
        writer.writeheader()
        for r in rows:
            writer.writerow({k: (r[k] if r[k] is not None else '') for k in CAMPOS})
        out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
        return send_file(out, mimetype='text/csv', as_attachment=True,
                         download_name='materiales.csv')

    elif formato == 'json':
        data = [{k: (r[k] if r[k] is not None else '') for k in CAMPOS} for r in rows]
        out = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2, default=str).encode('utf-8'))
        return send_file(out, mimetype='application/json', as_attachment=True,
                         download_name='materiales.json')

    elif formato == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Materiales'

        # Cabecera con estilo
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='2980B9')

        ws.append(CAMPOS)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for r in rows:
            ws.append([(r[k] if r[k] is not None else '') for k in CAMPOS])

        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='materiales.xlsx')

    return 'Formato no válido', 400


@materiales_bp.route('/materiales/plantilla/<formato>')
def plantilla(formato):
    tenant_id = get_tenant_filter()
    HEADERS = ['codigo', 'nombre', 'descripcion', 'codigo_barras',
               'categoria_id', 'stock_minimo', 'stock_maximo', 'unidad_medida_id', 'trazabilidad',
               'metodo_picking', 'peso_bruto', 'peso_neto', 'id_proveedor_habitual', 'codigo_referencia_prov']
    EJEMPLO = ['MAT001', 'Ejemplo Material', 'Descripción opcional', '',
                '1', '0.00', '100.00', '1', 'ninguna', 'libre', '0.500', '0.450', '1', 'REF-PROV-001']

    # Obtener categorias, unidades y proveedores
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("SELECT id_categoria, nombre FROM categorias WHERE activo = 1 AND (%s IS NULL OR tenant_id = %s) ORDER BY nombre", (tenant_id, tenant_id))
            categorias = cursor.fetchall()
            cursor.execute("SELECT id_unidad, nombre, simbolo FROM unidades_medida WHERE activo = 1 AND (%s IS NULL OR tenant_id = %s) ORDER BY nombre", (tenant_id, tenant_id))
            unidades = cursor.fetchall()
            cursor.execute("SELECT id, razonsocial FROM proveedores WHERE activo = 1 AND (%s IS NULL OR tenant_id = %s) ORDER BY razonsocial", (tenant_id, tenant_id))
            proveedores = cursor.fetchall()
    finally:
        conn.close()

    if formato == 'csv':
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(['# PLANTILLA MATERIALES'])
        writer.writerow(HEADERS)
        writer.writerow(EJEMPLO)
        writer.writerow([])
        writer.writerow(['# REFERENCIAS'])
        writer.writerow(['# CATEGORIAS (categoria_id)'])
        writer.writerow(['id_categoria', 'nombre'])
        for c in categorias:
            writer.writerow([c['id_categoria'], c['nombre']])
        writer.writerow([])
        writer.writerow(['# UNIDADES DE MEDIDA (unidad_medida_id)'])
        writer.writerow(['id_unidad', 'nombre', 'simbolo'])
        for u in unidades:
            writer.writerow([u['id_unidad'], u['nombre'], u['simbolo']])
        writer.writerow([])
        writer.writerow(['# PROVEEDORES (id_proveedor_habitual)'])
        writer.writerow(['id', 'razonsocial'])
        for p in proveedores:
            writer.writerow([p['id'], p['razonsocial']])
        out = io.BytesIO(buf.getvalue().encode('utf-8-sig'))
        return send_file(out, mimetype='text/csv', as_attachment=True,
                         download_name='plantilla_materiales.csv')

    elif formato == 'json':
        data = {
            'materiales': [dict(zip(HEADERS, EJEMPLO))],
            'categorias': [{'id_categoria': c['id_categoria'], 'nombre': c['nombre']} for c in categorias],
            'unidades': [{'id': u['id_unidad'], 'nombre': u['nombre'], 'abreviatura': u['simbolo']} for u in unidades],
            'proveedores': [{'id': p['id'], 'razonsocial': p['razonsocial']} for p in proveedores]
        }
        out = io.BytesIO(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
        return send_file(out, mimetype='application/json', as_attachment=True,
                         download_name='plantilla_materiales.json')

    elif formato == 'xlsx':
        wb = openpyxl.Workbook()
        
        # Hoja de Materiales
        ws_mat = wb.active
        ws_mat.title = 'Materiales'
        ws_mat.append(HEADERS)
        ws_mat.append(EJEMPLO)
        
        # Estilo para cabecera
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='2980B9')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws_mat[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Ancho de columnas
        for col in ws_mat.columns:
            ws_mat.column_dimensions[col[0].column_letter].width = 20
        
        # Hoja de Categorías
        ws_cat = wb.create_sheet('Categorias')
        ws_cat.append(['id_categoria', 'nombre'])
        for cell in ws_cat[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for c in categorias:
            ws_cat.append([c['id_categoria'], c['nombre']])
        for col in ws_cat.columns:
            ws_cat.column_dimensions[col[0].column_letter].width = 15
        
        # Hoja de Unidades
        ws_uni = wb.create_sheet('Unidades')
        ws_uni.append(['id_unidad', 'nombre', 'simbolo'])
        for cell in ws_uni[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for u in unidades:
            ws_uni.append([u['id_unidad'], u['nombre'], u['simbolo']])
        for col in ws_uni.columns:
            ws_uni.column_dimensions[col[0].column_letter].width = 15

        # Hoja de Proveedores
        ws_prov = wb.create_sheet('Proveedores')
        ws_prov.append(['id', 'razonsocial'])
        for cell in ws_prov[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for p in proveedores:
            ws_prov.append([p['id'], p['razonsocial']])
        for col in ws_prov.columns:
            ws_prov.column_dimensions[col[0].column_letter].width = 20

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return send_file(out,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='plantilla_materiales.xlsx')

    return 'Formato no válido', 400


@materiales_bp.route('/materiales/eliminar/<int:id>', methods=['POST'])
def eliminar(id):
    tenant_id = get_tenant_filter()
    conn = get_db_connection()
    try:
        with conn.cursor() as cursor:
            cursor.execute("DELETE FROM materiales WHERE id = %s AND (%s IS NULL OR tenant_id = %s)", (id, tenant_id, tenant_id))
            conn.commit()
            flash("Material eliminado", "success")
    finally:
        conn.close()
    return redirect(url_for('materiales.listar'))